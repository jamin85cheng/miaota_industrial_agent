"""
报表导出模块

功能需求: V-07 报表导出 - PDF/Excel导出
作者: Frontend Team + Data Team
"""

import json
from typing import Dict, List, Optional, Any
from datetime import datetime, timedelta
from pathlib import Path
from dataclasses import dataclass
from enum import Enum
import pandas as pd
from loguru import logger


class ReportType(Enum):
    """报表类型"""
    DAILY = "daily"           # 日报
    WEEKLY = "weekly"         # 周报
    MONTHLY = "monthly"       # 月报
    ALERT_SUMMARY = "alerts"  # 告警汇总
    DEVICE_HEALTH = "health"  # 设备健康
    CUSTOM = "custom"         # 自定义


@dataclass
class ReportData:
    """报表数据"""
    title: str
    period: str
    generated_at: datetime
    data: Dict[str, Any]
    charts: List[Dict]  # 图表数据


class ReportExporter:
    """
    报表导出器
    
    支持格式:
    - Excel (.xlsx)
    - PDF (.pdf)
    - CSV (.csv)
    - JSON (.json)
    """
    
    def __init__(self, output_dir: str = "data/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"报表导出器已初始化: {output_dir}")
    
    def export_excel(self, report_data: ReportData, filename: Optional[str] = None) -> str:
        """
        导出Excel报表
        
        Args:
            report_data: 报表数据
            filename: 文件名 (可选)
            
        Returns:
            导出文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import LineChart, Reference
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logger.error("缺少openpyxl依赖，请安装: pip install openpyxl")
            return ""
        
        if filename is None:
            filename = f"REPORT_{report_data.period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = self.output_dir / filename
        
        # 创建工作簿
        wb = Workbook()
        
        # 样式定义
        title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. 概览页
        ws_overview = wb.active
        ws_overview.title = "概览"
        
        # 标题
        ws_overview['A1'] = report_data.title
        ws_overview['A1'].font = title_font
        ws_overview['A1'].fill = title_fill
        ws_overview['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_overview.merge_cells('A1:E1')
        ws_overview.row_dimensions[1].height = 30
        
        # 基本信息
        ws_overview['A3'] = '报表周期'
        ws_overview['B3'] = report_data.period
        ws_overview['A4'] = '生成时间'
        ws_overview['B4'] = report_data.generated_at.strftime('%Y-%m-%d %H:%M:%S')
        
        # 关键指标
        row = 7
        ws_overview[f'A{row}'] = '关键指标'
        ws_overview[f'A{row}'].font = header_font
        ws_overview[f'A{row}'].fill = header_fill
        ws_overview.merge_cells(f'A{row}:E{row}')
        
        row += 1
        headers = ['指标名称', '数值', '单位', '环比', '状态']
        for col, header in enumerate(headers, 1):
            cell = ws_overview.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # 填充数据
        metrics = report_data.data.get('metrics', [])
        for metric in metrics:
            row += 1
            ws_overview.cell(row, 1, metric.get('name', ''))
            ws_overview.cell(row, 2, metric.get('value', ''))
            ws_overview.cell(row, 3, metric.get('unit', ''))
            ws_overview.cell(row, 4, metric.get('change', ''))
            status_cell = ws_overview.cell(row, 5, metric.get('status', ''))
            
            # 状态颜色
            if metric.get('status') == '正常':
                status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif metric.get('status') == '警告':
                status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif metric.get('status') == '异常':
                status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # 2. 详细数据页
        if 'details' in report_data.data:
            ws_details = wb.create_sheet("详细数据")
            
            # 表头
            headers = list(report_data.data['details'][0].keys()) if report_data.data['details'] else []
            for col, header in enumerate(headers, 1):
                cell = ws_details.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # 数据
            for row_idx, record in enumerate(report_data.data['details'], 2):
                for col_idx, header in enumerate(headers, 1):
                    ws_details.cell(row_idx, col_idx, record.get(header, ''))
        
        # 3. 告警页
        if 'alerts' in report_data.data:
            ws_alerts = wb.create_sheet("告警记录")
            
            alerts = report_data.data['alerts']
            if alerts:
                headers = list(alerts[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws_alerts.cell(1, col, header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                
                for row_idx, alert in enumerate(alerts, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws_alerts.cell(row_idx, col_idx, alert.get(header, ''))
        
        # 调整列宽
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存
        wb.save(filepath)
        logger.info(f"Excel报表已导出: {filepath}")
        
        return str(filepath)
    
    def export_csv(self, report_data: ReportData, filename: Optional[str] = None) -> str:
        """导出CSV报表"""
        if filename is None:
            filename = f"REPORT_{report_data.period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        filepath = self.output_dir / filename
        
        # 将详细数据转为DataFrame
        if 'details' in report_data.data:
            df = pd.DataFrame(report_data.data['details'])
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            logger.info(f"CSV报表已导出: {filepath}")
            return str(filepath)
        else:
            logger.warning("没有详细数据可导出")
            return ""
    
    def export_json(self, report_data: ReportData, filename: Optional[str] = None) -> str:
        """导出JSON报表"""
        if filename is None:
            filename = f"REPORT_{report_data.period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        filepath = self.output_dir / filename
        
        data = {
            'title': report_data.title,
            'period': report_data.period,
            'generated_at': report_data.generated_at.isoformat(),
            'data': report_data.data
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"JSON报表已导出: {filepath}")
        return str(filepath)
    
    def export_pdf(self, report_data: ReportData, filename: Optional[str] = None) -> str:
        """导出PDF报表"""
        # 复用诊断报告生成器的PDF功能
        try:
            from src.models.diagnosis_report import ReportGenerator
            
            generator = ReportGenerator(output_dir=str(self.output_dir))
            
            # 转换为Report格式
            from dataclasses import dataclass
            from src.models.diagnosis_report import DiagnosisReport
            
            report = DiagnosisReport(
                report_id=f"RPT_{datetime.now().strftime('%Y%m%d%H%M%S')}",
                diagnosis_id="",
                created_at=report_data.generated_at,
                device_name=report_data.title,
                symptoms=f"报表周期: {report_data.period}",
                root_cause="",
                confidence=0.0,
                possible_causes=[],
                suggested_actions=[],
                spare_parts=[],
                references=[],
                similar_cases=[],
                trend_charts=report_data.charts
            )
            
            return generator.generate_pdf(report)
            
        except Exception as e:
            logger.error(f"PDF导出失败: {e}")
            return ""


class ReportGeneratorUtil:
    """
    报表生成工具
    
    预置常用报表模板
    """
    
    def __init__(self):
        self.exporter = ReportExporter()
    
    def generate_daily_report(self, date: Optional[datetime] = None) -> str:
        """生成日报"""
        if date is None:
            date = datetime.now() - timedelta(days=1)
        
        report_data = ReportData(
            title="运营日报",
            period=date.strftime('%Y-%m-%d'),
            generated_at=datetime.now(),
            data={
                'metrics': [
                    {'name': '数据采集量', 'value': 86400, 'unit': '条', 'change': '+5%', 'status': '正常'},
                    {'name': '告警次数', 'value': 12, 'unit': '次', 'change': '-20%', 'status': '正常'},
                    {'name': '设备在线率', 'value': 98.5, 'unit': '%', 'change': '+0.5%', 'status': '正常'},
                    {'name': '异常检测数', 'value': 3, 'unit': '个', 'change': '0%', 'status': '警告'},
                ],
                'alerts': [
                    {'时间': '08:30', '级别': '警告', '描述': 'pH值偏高', '状态': '已处理'},
                    {'时间': '14:20', '级别': '紧急', '描述': '溶解氧过低', '状态': '已处理'},
                ]
            },
            charts=[]
        )
        
        return self.exporter.export_excel(report_data)
    
    def generate_alert_report(self, start_date: datetime, end_date: datetime) -> str:
        """生成告警汇总报表"""
        report_data = ReportData(
            title="告警汇总报表",
            period=f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}",
            generated_at=datetime.now(),
            data={
                'metrics': [
                    {'name': '总告警数', 'value': 45, 'unit': '次', 'change': '+10%', 'status': '警告'},
                    {'name': '紧急告警', 'value': 5, 'unit': '次', 'change': '-2', 'status': '正常'},
                    {'name': '平均响应时间', 'value': 8, 'unit': '分钟', 'change': '-20%', 'status': '正常'},
                    {'name': '未处理告警', 'value': 2, 'unit': '个', 'change': '+1', 'status': '异常'},
                ],
                'alerts': [
                    {'告警ID': 'ALT001', '时间': '2024-01-15 08:30', '级别': '警告', '规则': 'pH异常', '状态': '已确认'},
                    {'告警ID': 'ALT002', '时间': '2024-01-15 14:20', '级别': '紧急', '规则': '缺氧异常', '状态': '已确认'},
                ]
            },
            charts=[]
        )
        
        return self.exporter.export_excel(report_data)


# 使用示例
if __name__ == "__main__":
    # 创建报表数据
    report_data = ReportData(
        title="污水处理厂运营周报",
        period="2024-01-08 至 2024-01-14",
        generated_at=datetime.now(),
        data={
            'metrics': [
                {'name': '进水流量', 'value': 12500, 'unit': 'm³', 'change': '+3%', 'status': '正常'},
                {'name': 'COD去除率', 'value': 92.5, 'unit': '%', 'change': '+1.2%', 'status': '正常'},
                {'name': '氨氮去除率', 'value': 88.3, 'unit': '%', 'change': '-2%', 'status': '警告'},
                {'name': '设备完好率', 'value': 96.8, 'unit': '%', 'change': '0%', 'status': '正常'},
                {'name': '电耗', 'value': 0.28, 'unit': 'kWh/m³', 'change': '-5%', 'status': '正常'},
            ],
            'details': [
                {'日期': '2024-01-08', '进水量': 1800, 'COD': 45, '氨氮': 2.1, '状态': '正常'},
                {'日期': '2024-01-09', '进水量': 1750, 'COD': 42, '氨氮': 1.9, '状态': '正常'},
                {'日期': '2024-01-10', '进水量': 1820, 'COD': 48, '氨氮': 2.5, '状态': '警告'},
            ],
            'alerts': [
                {'时间': '2024-01-10 09:15', '级别': '警告', '描述': '1#曝气池DO偏低', '处理人': '张工'},
                {'时间': '2024-01-12 16:30', '级别': '紧急', '描述': '提升泵故障', '处理人': '李工'},
            ]
        },
        charts=[]
    )
    
    # 导出各种格式
    exporter = ReportExporter()
    
    excel_path = exporter.export_excel(report_data)
    print(f"Excel: {excel_path}")
    
    csv_path = exporter.export_csv(report_data)
    print(f"CSV: {csv_path}")
    
    json_path = exporter.export_json(report_data)
    print(f"JSON: {json_path}")
    
    # 生成预置报表
    generator = ReportGeneratorUtil()
    daily_path = generator.generate_daily_report()
    print(f"\n日报: {daily_path}")
